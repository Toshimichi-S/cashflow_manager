"""
CSV / Excel エクスポートユーティリティ
"""

from __future__ import annotations
import csv
import os
import shutil
from datetime import datetime
from pathlib import Path
from controllers.cashflow_controller import compute_annual_summaries, get_fiscal_months
from database.db_manager import DB_PATH


def export_annual_csv(fiscal_year: int, output_path: str) -> None:
    """年間収支サマリーをCSVに書き出す"""
    summaries = compute_annual_summaries(fiscal_year)
    months_label = get_fiscal_months(fiscal_year)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow([
            "年月", "売上（予定）", "売上（実績）", "売上差異",
            "経費（予定）", "経費（実績）", "経費差異",
            "収支（予定）", "収支（実績）", "収支差異",
        ])
        for s in summaries:
            writer.writerow([
                f"{s.year}年{s.month}月",
                int(s.sale_planned), int(s.sale_actual), int(s.sale_diff),
                int(s.expense_planned), int(s.expense_actual), int(s.expense_diff),
                int(s.profit_planned), int(s.profit_actual),
                int(s.profit_actual - s.profit_planned),
            ])


def export_annual_excel(fiscal_year: int, output_path: str) -> None:
    """年間収支サマリーをExcelに書き出す"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxlがインストールされていません: pip install openpyxl")

    summaries = compute_annual_summaries(fiscal_year)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{fiscal_year}年度 収支サマリー"

    # スタイル定義
    header_fill = PatternFill("solid", fgColor="185FA5")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    pos_font = Font(color="1D9E75", size=10)
    neg_font = Font(color="D85A30", size=10)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    headers = [
        "年月", "売上（予定）", "売上（実績）", "売上差異",
        "経費（予定）", "経費（実績）", "経費差異",
        "収支（予定）", "収支（実績）", "収支差異",
    ]
    ws.append(headers)
    for col, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    for i, s in enumerate(summaries, 2):
        row_data = [
            f"{s.year}年{s.month}月",
            int(s.sale_planned), int(s.sale_actual), int(s.sale_diff),
            int(s.expense_planned), int(s.expense_actual), int(s.expense_diff),
            int(s.profit_planned), int(s.profit_actual),
            int(s.profit_actual - s.profit_planned),
        ]
        ws.append(row_data)
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col)
            cell.border = thin
            if col == 1:
                cell.alignment = center
            elif isinstance(val, int):
                cell.alignment = right
                cell.number_format = "#,##0"
                if col in (4, 7, 10):  # 差異列
                    cell.font = pos_font if val >= 0 else neg_font
            # 偶数行に薄いグレー背景
            if i % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F5F5")

    # 列幅調整
    col_widths = [12, 14, 14, 12, 14, 14, 12, 14, 14, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)


def backup_database(dest_dir: str) -> str:
    """DBファイルをバックアップしてパスを返す"""
    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = dest / f"cashflow_backup_{timestamp}.db"
    shutil.copy2(DB_PATH, backup_path)
    return str(backup_path)


def restore_database(backup_path: str) -> None:
    """バックアップDBをリストアする（既存DBを上書き）"""
    shutil.copy2(backup_path, DB_PATH)
